# -*- coding: utf-8 -*-
"""
Utilidades para exportación de datos a Excel (XLSX).
Provee funciones helper para generar reportes descargables.
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
import structlog

logger = structlog.get_logger(__name__)


def crear_response_excel(filename, sheet_name='Datos'):
    """
    Crea HttpResponse configurado para descarga de Excel.
    
    Args:
        filename (str): Nombre del archivo (sin extensión)
        sheet_name (str): Nombre de la hoja (opcional)
    
    Returns:
        tuple: (response, BytesIO buffer, workbook)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl_no_instalado")
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Buffer para guardar en memoria
    buffer = BytesIO()
    
    # Crear response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_{timestamp}.xlsx"'
    
    return response, buffer, wb


def aplicar_estilos_header(ws, num_columns):
    """
    Aplica estilos profesionales a la fila de encabezados.
    
    Args:
        ws: Worksheet de openpyxl
        num_columns (int): Número de columnas
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_side = Side(style='thin', color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border


def ajustar_anchos_columnas(ws, min_width=10, max_width=50):
    """
    Ajusta automáticamente el ancho de las columnas según contenido.
    
    Args:
        ws: Worksheet de openpyxl
        min_width (int): Ancho mínimo
        max_width (int): Ancho máximo
    """
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        adjusted_width = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


def exportar_queryset_a_excel(queryset, filename, headers, campos, sheet_name='Datos'):
    """
    Exporta un QuerySet de Django a Excel.
    
    Args:
        queryset: QuerySet de Django
        filename (str): Nombre del archivo
        headers (list): Lista de nombres de columnas
        campos (list): Lista de campos del modelo o lambdas
        sheet_name (str): Nombre de la hoja
    
    Returns:
        HttpResponse: Archivo Excel para descarga
    
    Example:
        >>> from .models import Ticket
        >>> qs = Ticket.objects.all()
        >>> headers = ['UUID', 'Estado', 'Fecha Creación']
        >>> campos = ['uuid', 'estado', lambda t: t.created_at.strftime('%Y-%m-%d')]
        >>> return exportar_queryset_a_excel(qs, 'tickets', headers, campos)
    """
    logger.info("exportando_a_excel", filename=filename, registros=queryset.count())
    
    try:
        response, buffer, wb = crear_response_excel(filename, sheet_name)
        ws = wb.active
        
        # Escribir encabezados
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        aplicar_estilos_header(ws, len(headers))
        
        # Escribir datos
        for obj in queryset:
            row = []
            for campo in campos:
                if callable(campo):
                    # Lambda o función
                    valor = campo(obj)
                elif '.' in campo:
                    # Relación (ej: 'trabajador.nombre')
                    partes = campo.split('.')
                    valor = obj
                    for parte in partes:
                        valor = getattr(valor, parte, '')
                else:
                    # Campo directo
                    valor = getattr(obj, campo, '')
                
                # Formatear valores especiales
                if hasattr(valor, 'isoformat'):
                    valor = valor.isoformat()
                elif isinstance(valor, bool):
                    valor = 'Sí' if valor else 'No'
                elif valor is None:
                    valor = ''
                
                row.append(valor)
            
            ws.append(row)
        
        # Ajustar anchos
        ajustar_anchos_columnas(ws)
        
        # Guardar en buffer
        wb.save(buffer)
        buffer.seek(0)
        
        # Escribir en response
        response.write(buffer.getvalue())
        buffer.close()
        
        logger.info("exportacion_excel_exitosa", filename=filename)
        return response
        
    except Exception as e:
        logger.error("error_exportando_excel", error=str(e))
        raise


def exportar_dict_list_a_excel(data_list, filename, headers=None, sheet_name='Datos'):
    """
    Exporta una lista de diccionarios a Excel.
    
    Args:
        data_list (list): Lista de diccionarios
        filename (str): Nombre del archivo
        headers (list): Lista de encabezados (opcional, usa keys del primer dict)
        sheet_name (str): Nombre de la hoja
    
    Returns:
        HttpResponse: Archivo Excel para descarga
    
    Example:
        >>> data = [
        ...     {'nombre': 'Juan', 'edad': 30, 'ciudad': 'Santiago'},
        ...     {'nombre': 'María', 'edad': 25, 'ciudad': 'Valparaíso'},
        ... ]
        >>> return exportar_dict_list_a_excel(data, 'personas')
    """
    logger.info("exportando_dict_list_a_excel", filename=filename, registros=len(data_list))
    
    try:
        if not data_list:
            raise ValueError("La lista de datos está vacía")
        
        # Obtener headers si no se proveen
        if headers is None:
            headers = list(data_list[0].keys())
        
        response, buffer, wb = crear_response_excel(filename, sheet_name)
        ws = wb.active
        
        # Escribir encabezados
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        aplicar_estilos_header(ws, len(headers))
        
        # Escribir datos
        for item in data_list:
            row = []
            for header in headers:
                valor = item.get(header, '')
                
                # Formatear valores
                if hasattr(valor, 'isoformat'):
                    valor = valor.isoformat()
                elif isinstance(valor, bool):
                    valor = 'Sí' if valor else 'No'
                elif isinstance(valor, (list, dict)):
                    valor = str(valor)
                elif valor is None:
                    valor = ''
                
                row.append(valor)
            
            ws.append(row)
        
        # Ajustar anchos
        ajustar_anchos_columnas(ws)
        
        # Guardar en buffer
        wb.save(buffer)
        buffer.seek(0)
        
        # Escribir en response
        response.write(buffer.getvalue())
        buffer.close()
        
        logger.info("exportacion_dict_excel_exitosa", filename=filename)
        return response
        
    except Exception as e:
        logger.error("error_exportando_dict_excel", error=str(e))
        raise


def agregar_hoja_resumen(wb, titulo, estadisticas):
    """
    Agrega una hoja de resumen con estadísticas al workbook.
    
    Args:
        wb: Workbook de openpyxl
        titulo (str): Título de la hoja
        estadisticas (dict): Diccionario con estadísticas a mostrar
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return
    
    ws = wb.create_sheet(titulo)
    
    # Título
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A1:B1')
    
    # Estadísticas
    row = 3
    for key, value in estadisticas.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
