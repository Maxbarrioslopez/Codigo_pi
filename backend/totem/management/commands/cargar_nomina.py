"""
Management command para cargar nómina de trabajadores desde archivo CSV/Excel/JSON.

Uso:
    python manage.py cargar_nomina archivo.csv
    python manage.py cargar_nomina archivo.xlsx --sheet "Nomina"
    python manage.py cargar_nomina archivo.json

Formatos soportados:

CSV:
    rut,nombre,seccion,contrato,sucursal,beneficio,ciclo_id,observaciones,email,telefono
    12345678-9,Juan Pérez,Operaciones,INDEFINIDO,CASABLANCA,CAJA,8,,,+56912345678

Excel (.xlsx):
    Mismo formato que CSV, se lee de la hoja "Nomina" por defecto

JSON (nomina_estructura.json):
    {
        "ciclo": {
            "id": 8,
            "fecha_inicio": "2025-12-01",
            "fecha_fin": "2025-12-31",
            "nombre": "Diciembre 2025"
        },
        "trabajadores": [
            {
                "rut": "12345678-9",
                "nombre": "Juan Pérez",
                "seccion": "Operaciones",
                "contrato": "INDEFINIDO",
                "sucursal": "Casablanca",
                "beneficio": {
                    "tipo": "Caja",
                    "categoria": "Estándar",
                    "valor": null,
                    "descripcion": "Caja de mercadería",
                    "activo": true
                },
                "observaciones": "",
                "email": "",
                "telefono": ""
            }
        ],
        "metadata": {
            "total_trabajadores": 10,
            "total_con_beneficio": 7,
            "total_sin_beneficio": 3,
            "fecha_generacion": "2025-12-01T10:00:00Z"
        }
    }

Beneficio puede ser:
    - CAJA: Caja de mercadería
    - VALE: Vale de compra
    - VALE:5000: Vale con monto específico
    - MONTO:10000: Monto en dinero
    - NO / NINGUNO / SIN BENEFICIO / N/A: Sin beneficio (usar observaciones para explicar)

Observaciones: Campo opcional para explicar por qué no recibe beneficio
"""
# Nota: Uso orientado a desarrollo/operaciones manuales (dev-only).
# No se ejecuta en runtime de la aplicación.

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from totem.models import Trabajador, Sucursal
from totem.validators import RUTValidator, InputSanitizer
import csv
import json
import logging

logger = logging.getLogger(__name__)

# Conjuntos canonicos para estandarizar contratos y sucursales
CONTRATOS_CANONICOS = {
    'indefinido': 'Indefinido',
    'plazo fijo': 'Plazo Fijo',
    'plazo_fijo': 'Plazo Fijo',
    'plazo-fijo': 'Plazo Fijo',
    'plazofijo': 'Plazo Fijo',
    'part time': 'Part Time',
    'part_time': 'Part Time',
    'part-time': 'Part Time',
    'parttime': 'Part Time',
    'honorario': 'Honorarios',
    'honorarios': 'Honorarios',
    'externo': 'Externos',
    'externos': 'Externos',
}

SUCURSALES_CANONICAS = {
    'casablanca': 'Casablanca',
    'casa blanca': 'Casablanca',
    'valparaiso planta bif': 'Valparaiso Planta BIF',
    'valparaiso bif': 'Valparaiso Planta BIF',
    'bif': 'Valparaiso Planta BIF',
    'valparaiso planta bic': 'Valparaiso Planta BIC',
    'valparaiso bic': 'Valparaiso Planta BIC',
    'bic': 'Valparaiso Planta BIC',
}


class Command(BaseCommand):
    help = 'Carga nómina de trabajadores: rut, nombre, seccion, contrato, sucursal, beneficio, observaciones'
    
    # Columnas requeridas (beneficio ahora es opcional)
    COLUMNAS_REQUERIDAS = ['rut', 'nombre', 'seccion', 'contrato', 'sucursal']

    # Columnas opcionales
    COLUMNAS_OPCIONALES = ['beneficio', 'observaciones', 'email', 'telefono']

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta al archivo CSV o Excel (.xlsx)'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Nomina',
            help='Nombre de la hoja en archivo Excel (default: Nomina)'
        )
        parser.add_argument(
            '--actualizar',
            action='store_true',
            help='Actualizar trabajadores existentes (default: solo crear nuevos)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simular carga sin guardar en BD (para verificar)'
        )
        parser.add_argument(
            '--sucursal-defecto',
            type=str,
            help='Código de sucursal por defecto si no existe la especificada'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        actualizar = options['actualizar']
        dry_run = options['dry_run']
        sucursal_defecto = options.get('sucursal_defecto')
        
        if dry_run:
            self.stdout.write(self.style.WARNING('Modo DRY-RUN: No se guardará en BD'))
        
        self.stdout.write(f'Cargando archivo: {archivo}')
        
        # Detectar tipo de archivo
        if archivo.endswith('.csv'):
            trabajadores = self._cargar_csv(archivo)
        elif archivo.endswith(('.xlsx', '.xls')):
            trabajadores = self._cargar_excel(archivo, options['sheet'])
        elif archivo.endswith('.json'):
            trabajadores = self._cargar_json(archivo)
        else:
            raise CommandError('Formato no soportado. Use .csv, .xlsx o .json')
        
        if not trabajadores:
            raise CommandError('No se encontraron trabajadores válidos en el archivo')
        
        # Procesar trabajadores
        self._procesar_trabajadores(trabajadores, actualizar, dry_run, sucursal_defecto)

    def _cargar_csv(self, archivo):
        """Carga trabajadores desde archivo CSV."""
        trabajadores = []
        
        try:
            with open(archivo, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                if not reader.fieldnames:
                    raise CommandError('El archivo CSV está vacío o no tiene encabezados')
                
                # Validar columnas requeridas (case-insensitive)
                columnas_archivo = {c.lower().strip() for c in reader.fieldnames}
                columnas_faltantes = [c for c in self.COLUMNAS_REQUERIDAS if c not in columnas_archivo]
                if columnas_faltantes:
                    raise CommandError(
                        f'❌ Faltan columnas requeridas: {", ".join(columnas_faltantes)}\n'
                        f'📋 Columnas encontradas: {", ".join(columnas_archivo)}\n'
                        f'✅ Columnas requeridas: {", ".join(self.COLUMNAS_REQUERIDAS)}\n'
                        f'📝 Columnas opcionales: {", ".join(self.COLUMNAS_OPCIONALES)}'
                    )
                
                self.stdout.write(f'✓ Columnas validadas: {", ".join(columnas_archivo)}')
                
                for idx, row in enumerate(reader, start=2):  # start=2 (header es línea 1)
                    try:
                        # Normalizar keys a lowercase
                        row_lower = {k.lower().strip(): v for k, v in row.items()}
                        trabajador = self._parsear_fila(row_lower, idx)
                        if trabajador:
                            trabajadores.append(trabajador)
                    except Exception as e:
                        self.stdout.write(
                            self.style.ERROR(f'❌ Error línea {idx}: {e}')
                        )
                        continue
        
        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {archivo}')
        except Exception as e:
            raise CommandError(f'Error leyendo CSV: {e}')
        
        return trabajadores

    def _cargar_excel(self, archivo, sheet_name):
        """Carga trabajadores desde archivo Excel."""
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                'openpyxl no instalado. Instalar con: pip install openpyxl'
            )
        
        trabajadores = []
        
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True)
            
            # Obtener hoja
            ws = None
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                if ws:
                    self.stdout.write(
                        self.style.WARNING(
                            f'Hoja "{sheet_name}" no encontrada, usando: {ws.title}'
                        )
                    )
            
            if not ws:
                raise CommandError('No se pudo acceder a ninguna hoja del archivo Excel')
            
            # Leer headers
            headers = [str(cell.value).lower().strip() if cell.value else '' 
                      for cell in ws[1]]
            
            # Validar columnas
            columnas_faltantes = [c for c in self.COLUMNAS_REQUERIDAS if c not in headers]
            if columnas_faltantes:
                raise CommandError(
                    f'Faltan columnas requeridas: {", ".join(columnas_faltantes)}\n'
                    f'Columnas encontradas: {", ".join(headers)}'
                )
            
            # Procesar filas
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_dict = {headers[i]: (str(cell).strip() if cell is not None else '') 
                               for i, cell in enumerate(row) if i < len(headers)}
                    trabajador = self._parsear_fila(row_dict, idx)
                    if trabajador:
                        trabajadores.append(trabajador)
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f'Error fila {idx}: {e}')
                    )
                    continue
        
        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {archivo}')
        except Exception as e:
            raise CommandError(f'Error leyendo Excel: {e}')
        
        return trabajadores

    def _cargar_json(self, archivo):
        """Carga trabajadores desde archivo JSON (nomina_estructura.json).
        
        Formato esperado:
        {
            "ciclo": { "id": 8, ... },
            "trabajadores": [ { "rut": "...", "nombre": "...", ... } ],
            "metadata": { "total_trabajadores": 10, ... }
        }
        """
        trabajadores = []
        
        try:
            with open(archivo, 'r', encoding='utf-8-sig') as f:
                data = json.load(f)
            
            # Validar estructura JSON
            if 'trabajadores' not in data:
                raise CommandError(
                    'El archivo JSON debe contener una clave "trabajadores" con array de trabajadores'
                )
            
            if not isinstance(data['trabajadores'], list):
                raise CommandError(
                    'El campo "trabajadores" debe ser un array'
                )
            
            # Extraer ciclo_id del metadata del JSON si está disponible
            ciclo_id_json = None
            if 'ciclo' in data and isinstance(data['ciclo'], dict):
                ciclo_id_json = data['ciclo'].get('id')
                self.stdout.write(
                    self.style.SUCCESS(f'✓ Ciclo extraído del JSON: {ciclo_id_json}')
                )
            
            # Procesar cada trabajador del JSON
            for idx, trabajador_json in enumerate(data['trabajadores'], start=1):
                try:
                    # Convertir estructura JSON a formato compatible con _parsear_fila
                    row = {
                        'rut': str(trabajador_json.get('rut', '')).strip(),
                        'nombre': str(trabajador_json.get('nombre', '')).strip(),
                        'seccion': str(trabajador_json.get('seccion', '')).strip(),
                        'contrato': str(trabajador_json.get('contrato', '')).strip(),
                        'sucursal': str(trabajador_json.get('sucursal', '')).strip(),
                        'beneficio': str(trabajador_json.get('beneficio', {}).get('tipo', '')).strip(),
                        'ciclo_id': ciclo_id_json or trabajador_json.get('ciclo_id'),
                        'observaciones': str(trabajador_json.get('observaciones', '')).strip(),
                        'email': str(trabajador_json.get('email', '')).strip(),
                        'telefono': str(trabajador_json.get('telefono', '')).strip(),
                    }
                    
                    # Si el beneficio es un objeto complejo, extraer tipo
                    if isinstance(trabajador_json.get('beneficio'), dict):
                        beneficio_obj = trabajador_json['beneficio']
                        beneficio_tipo = beneficio_obj.get('tipo', '')
                        beneficio_valor = beneficio_obj.get('valor')
                        
                        # Reconstruir string de beneficio si hay valor
                        if beneficio_valor:
                            row['beneficio'] = f"{beneficio_tipo}:{beneficio_valor}"
                        else:
                            row['beneficio'] = beneficio_tipo
                    
                    trabajador = self._parsear_fila(row, idx)
                    if trabajador:
                        trabajadores.append(trabajador)
                
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f'Error procesando trabajador {idx} en JSON: {e}')
                    )
                    continue
        
        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {archivo}')
        except json.JSONDecodeError as e:
            raise CommandError(f'Error decodificando JSON: {e}')
        except Exception as e:
            raise CommandError(f'Error leyendo JSON: {e}')
        
        return trabajadores

    def _normalizar_contrato(self, contrato):
        """Devuelve el contrato canonico o None si no matchea."""
        if not contrato:
            return None
        key = ' '.join(contrato.lower().replace('_', ' ').replace('-', ' ').split())
        return CONTRATOS_CANONICOS.get(key)

    def _normalizar_sucursal(self, sucursal):
        """Devuelve la sucursal canonica o None si no matchea."""
        if not sucursal:
            return None
        key = ' '.join(sucursal.lower().split())
        return SUCURSALES_CANONICAS.get(key)

    def _parsear_fila(self, row, linea):
        """Parsea y valida una fila del archivo.
        
        Returns:
            dict con datos del trabajador o None si hay error crítico
        """
        try:
            # 1. VALIDAR Y LIMPIAR RUT
            rut = str(row.get('rut', '')).strip()
            if not rut:
                self.stdout.write(self.style.ERROR(f'Línea {linea}: RUT vacío'))
                return None
            
            rut = InputSanitizer.sanitize_rut(rut)
            rut_limpio = RUTValidator.limpiar_rut(rut)
            es_valido, error = RUTValidator.validar_formato(rut_limpio)
            
            if not es_valido:
                self.stdout.write(
                    self.style.ERROR(f'Línea {linea}: RUT inválido {rut} - {error}')
                )
                return None
            
            # 2. VALIDAR NOMBRE
            nombre = str(row.get('nombre', '')).strip()
            if not nombre:
                self.stdout.write(self.style.ERROR(f'Línea {linea}: Nombre vacío'))
                return None
            
            nombre = InputSanitizer.sanitize_string(nombre, max_length=200)
            
            # 3. CAMPOS OBLIGATORIOS (almacenamos todo en beneficio_disponible)
            seccion = InputSanitizer.sanitize_string(
                str(row.get('seccion', '')).strip(), max_length=100
            )
            contrato_raw = InputSanitizer.sanitize_string(
                str(row.get('contrato', '')).strip(), max_length=50
            )
            sucursal_raw = InputSanitizer.sanitize_string(
                str(row.get('sucursal', '')).strip(), max_length=50
            )

            contrato = self._normalizar_contrato(contrato_raw)
            if contrato_raw and not contrato:
                contrato = contrato_raw
                self.stdout.write(
                    self.style.WARNING(
                        f'Línea {linea}: Tipo de contrato no reconocido "{contrato_raw}". Use uno de: Indefinido, Plazo Fijo, Part Time, Honorarios, Externos.'
                    )
                )

            sucursal_codigo = self._normalizar_sucursal(sucursal_raw)
            if sucursal_raw and not sucursal_codigo:
                sucursal_codigo = sucursal_raw
                self.stdout.write(
                    self.style.WARNING(
                        f'Línea {linea}: Sucursal no reconocida "{sucursal_raw}". Use una de: Casablanca, Valparaiso Planta BIF, Valparaiso Planta BIC.'
                    )
                )
            
            # CICLO ID - REQUERIDO para asociar beneficio a ciclo específico
            ciclo_id = row.get('ciclo_id', '')
            if ciclo_id:
                try:
                    ciclo_id = int(ciclo_id)
                except (ValueError, TypeError):
                    self.stdout.write(
                        self.style.WARNING(f'Línea {linea}: ciclo_id inválido "{ciclo_id}" para {nombre}. Se usará None.')
                    )
                    ciclo_id = None
            else:
                ciclo_id = None
            
            if not seccion:
                self.stdout.write(
                    self.style.WARNING(f'Línea {linea}: Sección vacía para {nombre}')
                )
            
            if not contrato:
                self.stdout.write(
                    self.style.WARNING(f'Línea {linea}: Tipo de contrato vacío para {nombre}')
                )
            
            # 4. CAMPOS OPCIONALES
            email = row.get('email', '')
            if email:
                email = InputSanitizer.sanitize_email(str(email).strip())
            
            telefono = row.get('telefono', '')
            if telefono:
                telefono = InputSanitizer.sanitize_phone(str(telefono).strip())
            
            observaciones = InputSanitizer.sanitize_string(
                str(row.get('observaciones', '')).strip(), max_length=500
            )
            
            # 5. PROCESAR BENEFICIO
            beneficio_raw = str(row.get('beneficio', '')).strip().upper() if 'beneficio' in row else ''
            beneficio = {}
            sin_beneficio = False
            motivo_sin_beneficio = None
            
            if not beneficio_raw or beneficio_raw in ['NO', 'NINGUNO', 'SIN BENEFICIO', 'N/A', '-', 'NA']:
                sin_beneficio = True
                motivo_sin_beneficio = observaciones or 'No especificado en nómina'
                self.stdout.write(
                    self.style.NOTICE(
                        f'ℹ️  Línea {linea}: {nombre} - SIN BENEFICIO ({motivo_sin_beneficio})'
                    )
                )
            else:
                # Parsear beneficio: puede ser "CAJA", "VALE", "VALE:5000", "MONTO:10000"
                try:
                    if ':' in beneficio_raw:
                        tipo, valor = beneficio_raw.split(':', 1)
                        beneficio = {
                            'tipo': tipo.strip(),
                            'valor': int(valor) if tipo.strip() == 'MONTO' else valor.strip()
                        }
                    else:
                        beneficio = {
                            'tipo': beneficio_raw,
                            'valor': None
                        }
                    
                    # Agregar metadata adicional al beneficio
                    if seccion:
                        beneficio['seccion'] = seccion
                    if contrato:
                        beneficio['tipo_contrato'] = contrato
                    if email:
                        beneficio['email'] = email
                    if telefono:
                        beneficio['telefono'] = telefono
                    if sucursal_codigo:
                        beneficio['sucursal'] = sucursal_codigo
                    if ciclo_id:
                        beneficio['ciclo_id'] = ciclo_id
                    
                except Exception as e:
                    self.stdout.write(
                        self.style.WARNING(
                            f'Línea {linea}: Formato de beneficio inválido "{beneficio_raw}": {e}'
                        )
                    )
                    beneficio = {'tipo': beneficio_raw, 'error': str(e)}
            
            # Si no tiene beneficio, guardar metadata en beneficio_disponible de todos modos
            if sin_beneficio:
                beneficio = {
                    'tipo': 'SIN_BENEFICIO',
                    'motivo': motivo_sin_beneficio,
                    'seccion': seccion,
                    'tipo_contrato': contrato,
                    'email': email,
                    'telefono': telefono,
                    'sucursal': sucursal_codigo,
                }
                if ciclo_id:
                    beneficio['ciclo_id'] = ciclo_id
            
            # 6. CONSTRUIR OBJETO TRABAJADOR
            return {
                'rut': rut_limpio,
                'nombre': nombre,
                'beneficio_disponible': beneficio if beneficio else {},
                'sucursal_codigo': sucursal_codigo,
                '_sin_beneficio': sin_beneficio,
                '_motivo_sin_beneficio': motivo_sin_beneficio,
            }
            
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Línea {linea}: Error parseando fila: {e}')
            )
            logger.error(f'Error parseando línea {linea}: {row} - {e}', exc_info=True)
            return None

    @transaction.atomic
    def _procesar_trabajadores(self, trabajadores, actualizar, dry_run, sucursal_defecto=None):
        """Procesa y guarda trabajadores en BD.
        
        Args:
            trabajadores: Lista de diccionarios con datos de trabajadores
            actualizar: Si actualizar trabajadores existentes
            dry_run: Si simular sin guardar
            sucursal_defecto: Código de sucursal por defecto
        """
        creados = 0
        actualizados = 0
        errores = 0
        sin_beneficio_count = 0
        
        self.stdout.write(f'\nProcesando {len(trabajadores)} trabajadores...\n')
        
        for idx, data in enumerate(trabajadores, start=1):
            try:
                rut = data['rut']
                sucursal_codigo = data.pop('sucursal_codigo', None)
                sin_beneficio = data.pop('_sin_beneficio', False)
                motivo_sin_beneficio = data.pop('_motivo_sin_beneficio', None)
                observaciones = data.pop('observaciones', None)
                
                if sin_beneficio:
                    sin_beneficio_count += 1
                
                # Resolver sucursal
                sucursal_obj = None
                if sucursal_codigo:
                    sucursal_obj = Sucursal.objects.filter(codigo=sucursal_codigo).first()
                    if not sucursal_obj and sucursal_defecto:
                        sucursal_obj = Sucursal.objects.filter(codigo=sucursal_defecto).first()
                        if sucursal_obj:
                            self.stdout.write(
                                self.style.NOTICE(
                                    f'  {data["nombre"]}: Usando sucursal por defecto {sucursal_defecto}'
                                )
                            )
                    
                    if not sucursal_obj:
                        self.stdout.write(
                                self.style.WARNING(
                                    f'  {data["nombre"]}: Sucursal "{sucursal_codigo}" no existe'
                                )
                        )
                
                # Verificar si existe
                trabajador_existente = Trabajador.objects.filter(rut=rut).first()
                
                if trabajador_existente:
                    if actualizar:
                        if not dry_run:
                            # Actualizar solo campos que existen en el modelo
                            trabajador_existente.nombre = data['nombre']
                            trabajador_existente.beneficio_disponible = data.get('beneficio_disponible', {})
                            trabajador_existente.save()
                        
                        actualizados += 1
                        estado = 'SIN' if sin_beneficio else 'ACT'
                        msg = f'  {estado} {rut} - {data["nombre"]} (actualizado)'
                        if sin_beneficio:
                            msg += f' - SIN BENEFICIO: {motivo_sin_beneficio}'
                        self.stdout.write(self.style.WARNING(msg))
                    else:
                        self.stdout.write(
                            self.style.NOTICE(f'  {rut} - {data["nombre"]} (ya existe, omitido)')
                        )
                else:
                    if not dry_run:
                        # Crear nuevo trabajador con solo los campos del modelo
                        Trabajador.objects.create(
                            rut=data['rut'],
                            nombre=data['nombre'],
                            beneficio_disponible=data.get('beneficio_disponible', {})
                        )
                    
                    creados += 1
                    estado = 'SIN' if sin_beneficio else 'OK'
                    msg = f'  {estado} {rut} - {data["nombre"]} (creado)'
                    if sin_beneficio:
                        msg += f' - SIN BENEFICIO: {motivo_sin_beneficio}'
                    self.stdout.write(self.style.SUCCESS(msg))
            
            except Exception as e:
                errores += 1
                self.stdout.write(
                    self.style.ERROR(f'  Error procesando {data.get("rut", "UNKNOWN")}: {e}')
                )
                logger.error(f'Error cargando trabajador {data}: {e}', exc_info=True)
        
        # Resumen
        self.stdout.write('\n' + '=' * 70)
        if dry_run:
            self.stdout.write(self.style.WARNING('SIMULACIÓN (DRY-RUN) - No se guardó en BD'))
        else:
            self.stdout.write(self.style.SUCCESS('CARGA COMPLETADA'))
        
        self.stdout.write('=' * 70)
        self.stdout.write(f'Total procesados:           {len(trabajadores)}')
        self.stdout.write(self.style.SUCCESS(f'  ✓ Creados:                   {creados}'))
        
        if actualizar:
            self.stdout.write(self.style.WARNING(f'  Actualizados:              {actualizados}'))
        
        if sin_beneficio_count > 0:
            self.stdout.write(self.style.NOTICE(f'  Sin beneficio:            {sin_beneficio_count}'))
        
        if errores > 0:
            self.stdout.write(self.style.ERROR(f'  Errores:                   {errores}'))
        
        self.stdout.write('=' * 70)
        
        if sin_beneficio_count > 0:
            self.stdout.write('\nLos trabajadores sin beneficio fueron cargados correctamente.')
            self.stdout.write('   Revisa las observaciones en la base de datos para conocer el motivo.')
        
        if not dry_run and (creados + actualizados) > 0:
            self.stdout.write('\nTrabajadores cargados exitosamente en el sistema')
            self.stdout.write('   Los que tienen beneficio pueden generar tickets en el tótem')

        # Crear sucursales por defecto si no existen
        for nombre, codigo in [
            ('Casablanca', 'CASA'),
            ('Valparaiso Planta BIF', 'BIF'),
            ('Valparaiso Planta BIC', 'BIC')
        ]:
            sucursal, creado = Sucursal.objects.get_or_create(nombre=nombre, codigo=codigo)
            if creado:
                self.stdout.write(self.style.SUCCESS(f'Sucursal creada: {nombre} ({codigo})'))
            else:
                self.stdout.write(self.style.NOTICE(f'Sucursal ya existe: {nombre} ({codigo})'))

